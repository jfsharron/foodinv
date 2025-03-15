"""
================================================================================
 Program:           foodinv.py
 Software Engineer: Jonas Sharron
 Date:              15-January-2024

 Purpose:   This program will provides a user interface for a MYSQL database
            (foodinv) and generates a csv file to facilitate printing labels. 
================================================================================
"""

import sys
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import mysql.connector
from mysql.connector import Error
import functools
import xlsxwriter
import pandas.io.sql as sql
import numpy as np
from numpy import loadtxt
from prettytable import from_db_cursor
from prettytable import PrettyTable
import datetime
#from datetime import datetime
from datetime import date
import os
from termcolor import colored, cprint 
from colorama import Fore, Back, Style 
from tabulate import tabulate
import fpdf
import colorama
from colorama import Fore, Back, Style
import getopt
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch
import webbrowser
import time
from pretty_html_table import build_table
from sqlalchemy import create_engine
import inquirer
from copy import copy

# ==============================================================================
# establish database connection
# ==============================================================================

XUSER       = 'jfsharron'
XWORD       = 'marie151414'
HOST        = '192.168.2.107'
DATABASE    = 'foodinv'

try:
    CONNECTION = mysql.connector.connect(user=XUSER, password=XWORD,
    host=HOST, database=DATABASE)
    if CONNECTION.is_connected():
        db_Info = CONNECTION.get_server_info()
        print("Connected to MySQL Server version ", db_Info)
        global cursor
        cursor = CONNECTION.cursor()
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("You're connected to database: ", record)
except Error as e:
    print("Error while connecting to MySQL", e)

# ==============================================================================
# intialize lists and variables
# ==============================================================================    

type_list           = []
weight_unit_list    = []
sub_type_list       = []
LABELFILE           = '/data/share/foodinv/food_label.xlsx'
RUN_ONCE            = 1



# ==============================================================================
# user functions
# ==============================================================================    

def menu():
    """
    ============================================================================
    Function:       menu()
    Purpose:        entry point to allow user interaction with program
    Parameter(s):   -None-
    Return:         users desired action
    ============================================================================
    """
    os.system('cls')
    # main menu
    #----------
    goAgain = 1

    while goAgain == 1:
        # format screen
        # --------------
        now = datetime.datetime.now()
        print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
        print(("foodinv").rjust(80))
        print("-----------------------".rjust(80))
        print(Style.RESET_ALL)
        print('')
        print(Fore.GREEN + 'MAIN MENU')
        print(Fore.GREEN + '-------------------')
        print(Style.RESET_ALL)
        print('1\tINPUT NEW RECORD')
        print('2\tDISCARD MENU')
        print('3\tREPORTS')
        print('')
        print('')
        print('')
        print(Fore.RED + '0\tEXIT')
        print(Style.RESET_ALL)
        print('')
        print('')      

        menuOption = input("selection: ")

        if menuOption == '1':
            get_selections()
        elif menuOption == '2':
            discard()
        elif menuOption == '3':
            reportMenu() 
        elif menuOption == '0':    
            goAgain = 0 

        os.system('cls')   

def reportMenu():
    """
    ============================================================================
    Function:       reportMenu()
    Purpose:        generate reports
    Parameter(s):   -None-
    Return:         users desired report
    ============================================================================
    """
    os.system('cls')
    # main menu
    #----------
    goAgain = 1

    while goAgain == 1:
        # format screen
        # --------------
        now = datetime.datetime.now()
        print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
        print(("foodinv").rjust(80))
        print("-----------------------".rjust(80))
        print(Style.RESET_ALL)
        print('')
        print(Fore.GREEN + 'REPORT MENU')
        print(Fore.GREEN + '-------------------')
        print(Style.RESET_ALL)

        # menu options
        # -------------
        print('1\tREPORT INDEX')
        print('2\tRUN REPORT')
        print('3\t')
        print('4\t')
        print('5\t')
        print('')
        print('')
        print('')
        print(Fore.RED + '0\tEXIT')
        print(Style.RESET_ALL)
        print('')
        print('')      

        menuOption = input("selection: ")

        if menuOption == '1':
            rep_name = str(11111) 
            report_temp2(rep_name)          
        elif menuOption == '2':
            rep_name = input("Enter report name: ")
            report_temp2(rep_name)
        elif menuOption == '0':    
            goAgain = 0 

        os.system('cls') 

def report_temp2(rep_name):
  """
  =======================================================================
  Function:       report_temp2(rep_name)
  Purpose:        pull required fields from MySQL to generate report
  Parameter(s):   rep_name
  Return:         rep_name, rep_query, rep_no
  =======================================================================
  """
  cursor.execute("SELECT query FROM report WHERE name = " + rep_name)
  row = cursor.fetchone()
  for row in cursor:
      row = cursor.fetchone()
      row = str(row)    
  rep_query = row
  x = str(rep_query)
  x= x.strip("()")
  x = x.strip("'")
  x = x.rstrip("',")
  cursor.execute("SELECT description FROM report WHERE name = " + rep_name)
  row = cursor.fetchone()
  for row in cursor:
      row = cursor.fetchone()
      row = str(row)    
  rep_na = row
  y = str(rep_na)
  y= y.strip("()")
  y = y.strip("'")
  y = y.rstrip("',")
  report_temp(y, x, rep_name)

def report_temp(rep_name, rep_query, rep_no):
  """
  =======================================================================
  Function:       report_temp(rep_name, rep_query, rep_no)
  Purpose:        generate a report
  Parameter(s):   rep_name, rep_query, rep_no
  Return:         generated report
  =======================================================================
  """
  # format screen
  # --------------
  now = datetime.datetime.now()
  print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
  print(("foodinv").rjust(80))
  print("-----------------------".rjust(80))
  print(Style.RESET_ALL)
  print('')
  print(Fore.GREEN + rep_name + " " + rep_no)
  print(Fore.GREEN + '-------------------')
  print(Style.RESET_ALL)
  # display report on screen
  # ------------------------
  mysql_select_query = rep_query
  cursor = CONNECTION.cursor(buffered = True)
  cursor.execute(mysql_select_query)    
  mytable = from_db_cursor(cursor)
  mytable.align = "l"
  print(mytable)
  print('')
  # send report to browser
  # -----------------------
  printRep = input(Fore.YELLOW + 'To send this report to the browser '
                  'for printing or saving enter b or B, otherwise press '
                  'enter to return: ')
  print(Style.RESET_ALL)
  if printRep == "b" or printRep == "B":
      # generate data for report
      # ------------------------
      mysql_search_query = rep_query
      cursor = CONNECTION.cursor(buffered = True)
      cursor.execute(mysql_search_query)
      mytable1 = pd.read_sql(rep_query, CONNECTION)
      pd.set_option('display.expand_frame_repr', False)
      mytable2 = build_table(mytable1,
                           'grey_light',
                           font_size = 'small',
                           font_family = 'Open Sans, courier',
                           text_align = 'left ')
      
      # generate html content
      # ---------------------
      report = rep_name + "_" + rep_no
      html_content = f"<html> \
                      <head> <h2> FOODINV Records Report - {rep_name} {rep_no}\
                      </h2> \
                      <h3> <script>\
                      var timestamp = Date.now();\
                      var d = new Date(timestamp);\
                      document.write(d);\
                      </script>\
                      </h3>\
                      </head> \
                      <body> {mytable2} \
                      </body> \
                      </html>"
      with open('/data/share/foodinv/report/report_' + report +'.html', "w") \
          as html_file:
          html_file.write(html_content)
          print("Created")
      time.sleep(2)
      # display in browser
      # ------------------
      webbrowser.get(using='lynx').open \
          ("/data/share/foodinv/report/report_" + report + ".html")
      print('')
      print("You may also access your report from the Reports Directory")
      print('')
      wait = input("Press ENTER to return") 

def get_selections():
    """
    ============================================================================
    Function:       get_selections()
    Purpose:        allows entry of a new record into database
    Parameter(s):   -None-
    Return:         users desired action
    ============================================================================
    """
    current_date    = str(datetime.datetime.now())

    global code_no

    # get intial code_no
    # -------------------
    mysql_select_query = ("SELECT value FROM xcounter WHERE counter_id = 1")
    cursor = CONNECTION.cursor(buffered = True)
    cursor.execute(mysql_select_query) 
    code_no = cursor.fetchone()
    code_no = str(code_no)
    code_no = code_no.strip(",()'")
    global init_code_no
    init_code_no = code_no

    # type selection
    # ---------------  
    mysql_select_query = ("SELECT type FROM type")

    cursor = CONNECTION.cursor(buffered = True)
    cursor.execute(mysql_select_query)    
    rows = cursor.fetchall()
    for row in rows:
        row = str(row)
        row = row.strip(",()'")
        type_list.append(row)

    print('')
    questions = [
      inquirer.List('ptype',
                    message="What is the type?",
                    choices=type_list,
                    ),
    ]
    answers = inquirer.prompt(questions)
    global ptype 
    ptype = str(answers["ptype"])
    ptype = ptype.strip(",()'")

    # retrieve type_prefix
    # ---------------------
    qptype = ("'" + ptype + "'")
    mysql_select_query = ("SELECT type_prefix FROM type WHERE type = " + qptype)

    cursor = CONNECTION.cursor(buffered = True)
    cursor.execute(mysql_select_query) 
    global type_prefix
    type_prefix = cursor.fetchone()
    type_prefix = str(type_prefix)
    type_prefix = type_prefix.strip(",()'")

    # sub_type selection
    # -------------------
    # empty sub_type list and concatenate ptype value to pull choices from
    # the proper sub_type list
    # -------------------------
    sub_type_list = []
    sub = (ptype + "_sub")
    print(sub)
    mysql_select_sub_query = ("SELECT type FROM " + sub)

    # execute query to display choices
    # ---------------------------------
    cursor = CONNECTION.cursor(buffered = True)
    cursor.execute(mysql_select_sub_query)    
    rows = cursor.fetchall()
    for row in rows:
        row = str(row)
        row = row.strip(",()'")
        sub_type_list.append(row)

    questions = [
      inquirer.List('sub_type',
                    message="What is the sub_type?",
                    choices=sub_type_list,
                    ),
    ]
    answers = inquirer.prompt(questions)
    global sub_type 
    sub_type = str(answers["sub_type"])
    sub_type = sub_type.strip(",()'")

    # enter (optional) description
    # -----------------------------
    global description
    description = input("enter (optional) description: ")

    # enter weight 
    # -------------
    print('')
    global weight
    weight= input("enter net weight: ")
    
    # weight_unit selection
    # ----------------------  
    mysql_select_query = ("SELECT weight_unit FROM weight_unit_sub")

    cursor = CONNECTION.cursor(buffered = True)
    cursor.execute(mysql_select_query)    
    rows = cursor.fetchall()
    for row in rows:
        row = str(row)
        row = row.strip(",()'")
        weight_unit_list.append(row)

    print('')
    questions = [
      inquirer.List('weight_unit',
                    message="What is the weight unit?",
                    choices=weight_unit_list,
                    ),
    ]
    answers = inquirer.prompt(questions)
    global weight_unit  
    weight_unit = str(answers["weight_unit"])
    weight_unit = weight_unit.strip(",()'")

    # enter piece count
    # -----------------   
    global piece
    piece= input("enter number of pieces: ")

    # enter date 
    # -----------
    current_date = datetime.date.today()

    global pack_date
    print('')
    pack_date = input("enter date packaged (YYYY-MM-DD) or press enter for" + \
                       " current date: ") or current_date
    pack_date = str(pack_date)
    print('')

    # calculate code 
    # ---------------
    global qcode
    qcode = (type_prefix + code_no)
    code_no = int(code_no)
    code_no += 1

    # process selections
    # -------------------
    ver_pro()

def ver_pro():
    """
    ============================================================================
    Function:       ver_pro()
    Purpose:        allows user to verify input and proceed
    Parameter(s):   values from get_selections()
    Return:         users desired action
    ============================================================================
    """
    print("These are your values, are the correct?")
    print('')
    print("type: " + ptype)
    print("type_prefix: " + type_prefix)
    print("sub_type: " + sub_type)
    print("description: " + description)
    print("weight: " + weight)
    print("weight_unit: " + weight_unit)
    print("pieces: " + piece)
    print("pack_date: " + pack_date)
    print("qcode: " + qcode) 
    print('')  
    print(Fore.YELLOW + "Select 1 to save and return or press enter")
    print("Select 2 to NOT save and return")
    print("Select 3 to save and go to the main menu")
    print("Select 4 to NOT save and return to the main menu")
    print(Style.RESET_ALL)
    print("")
    verify = input("Selection: ") or 1

    if verify == '1':
        savereturn_query()
    elif verify == '2':
        get_selections()
    elif verify == '3':
        save_query() 
    elif verify == '4':  
        menu()
    
def save_query():
    """
    ============================================================================
    Function:       save_query()
    Purpose:        write data to MYSQL
    Parameter(s):   values from ver_pro()
    Return:         data written to MYSQL foodinv and return to main menu
    ============================================================================
    """ 
    data = (ptype, sub_type, description, weight, weight_unit, piece, pack_date,\
            qcode, 0)
    save_query = ("INSERT INTO inv (type, sub_type, description, net_weight, \
                  weight_unit, pieces, date_packaged, code, discard)"
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)")

    cursor = CONNECTION.cursor()
    cursor.execute(save_query, data)
    CONNECTION.commit()      

    up_data = (code_no, 1)
    code_update_query = ("UPDATE xcounter SET value = (%s) WHERE \
                         counter_id = (%s)")
    cursor = CONNECTION.cursor()
    cursor.execute(code_update_query, up_data)
    CONNECTION.commit() 

    spreadsheet(qcode)    

def savereturn_query():
    """
    ============================================================================
    Function:       savereturn_query()
    Purpose:        write data to MYSQL
    Parameter(s):   values from ver_pro()
    Return:         data written to MYSQL foodinv and return to input menu
    ============================================================================
    """ 

    data = (ptype, sub_type, description, weight, weight_unit, piece, pack_date,\
            qcode, 0)
    save_query = ("INSERT INTO inv (type, sub_type, description, net_weight, \
                  weight_unit, pieces, date_packaged, code, discard)"
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)")

    cursor = CONNECTION.cursor()
    cursor.execute(save_query, data)
    CONNECTION.commit()

    up_data = (code_no, 1)
    code_update_query = ("UPDATE xcounter SET value = (%s) WHERE \
                         counter_id = (%s)")
    cursor = CONNECTION.cursor()
    cursor.execute(code_update_query, up_data)
    CONNECTION.commit()  

    spreadsheet(qcode)     
    
    get_selections()

def spreadsheet(qcode):
    """
    ============================================================================
    Function:       spreadsheet(qcode)
    Purpose:        send records to spreadsheet for label printing
    Parameter(s):   values from save queries
    Return:         data sent to LABELFILE 
    ============================================================================
    """
    # create new worksheet and input data
    # ------------------------------------ 
    global RUN_ONCE
    if RUN_ONCE == 1:

        # create new worksheet
        # ---------------------
        current_date = str(datetime.date.today())
        sht_name = (current_date + "  " + str(init_code_no))

        workbook = load_workbook(LABELFILE)
        source_ws_name = 'template'
        target_ws_name = sht_name
        source_ws = workbook[source_ws_name]
        target_ws = copy(source_ws)
        target_ws.title = target_ws_name
        workbook._add_sheet(target_ws)

        workbook.save(LABELFILE)

        up_data = (sht_name, 1)
        sht_update_query = ("UPDATE xcounter SET sht_name = (%s) WHERE \
                             counter_id = (%s)")
        cursor = CONNECTION.cursor()
        cursor.execute(sht_update_query, up_data)
        CONNECTION.commit()

        RUN_ONCE = 0

    # input data
    # -----------   
    append_q = ("'" + qcode + "'")

    mysql_select_query = ("SELECT sht_name FROM xcounter WHERE counter_id = 1")
    cursor = CONNECTION.cursor(buffered = True)
    cursor.execute(mysql_select_query) 
    sht_name = cursor.fetchone()
    sht_name = str(sht_name)
    sht_name = sht_name.strip(",()'")

    df = sql.read_sql(("SELECT * FROM inv WHERE code = " + append_q), \
                       CONNECTION)

    # Load the existing workbook
    # --------------------------
    workbook_path = LABELFILE
    workbook = load_workbook(workbook_path)

    # Get the target worksheet
    # ------------------------
    target_ws_name = sht_name 
    target_ws = workbook[target_ws_name]

    # Get the maximum row number in the target worksheet
    # --------------------------------------------------
    max_row = target_ws.max_row + 1

    # Write the new row of data to the worksheet
    # ------------------------------------------
    for index, row in df.iterrows():
        for col_num, value in enumerate(row, start=1):
            target_ws.cell(row=max_row, column=col_num, value=value)

    # Save the modified workbook
    # --------------------------
    workbook.save(workbook_path)

def discard():
    """
    ============================================================================
    Function:       discard()
    Purpose:        mark discard field in database file as TRUE for used product
    Parameter(s):   -None-
    Return:         changes field in MYSQL database
    ============================================================================
    """    
    # discard menu
    # -------------

    # format screen
    # --------------
    os.system('cls')
    now = datetime.datetime.now()
    print(Fore.GREEN + now.strftime("%Y-%m-%d %H:%M:%S").rjust(80))
    print(("foodinv").rjust(80))
    print("-----------------------".rjust(80))
    print(Style.RESET_ALL)
    print('')
    print(Fore.GREEN + 'DISCARD MENU')
    print(Fore.GREEN + '-------------------')
    print(Style.RESET_ALL)

    discardList = ["manually", "batch"]

    questions = [
      inquirer.List('discardSel',
                    message="How do you want to enter data?",
                    choices = discardList,
                    ),
    ]

    answers = inquirer.prompt(questions)
    discardOption = str(answers["discardSel"])

    # discard operation
    # ------------------
    if discardOption == 'manually':
        ent_code = input("Please enter the qcode you wish to mark as discard: ")
        print("selection was manually")

        append_q = ("'" + ent_code + "'")
        append_q =str(append_q)

        mysql_select_query = ("SELECT * FROM inv WHERE code = " + append_q)
        cursor = CONNECTION.cursor(buffered = True)
        cursor.execute(mysql_select_query) 
        sel_res = cursor.fetchone()

        if sel_res == None:
            programPause = input("Record not found, press the <ENTER> "
                                 "key to continue...")

        else:
            sql_up_query = ("""UPDATE inv SET discard = 1 WHERE code = """ \
                             + append_q)
            cursor.execute(sql_up_query)
            CONNECTION.commit()

    elif discardOption == 'batch':
        # Load the workbook
        # -----------------
        workbook = openpyxl.load_workbook(LABELFILE)
        sheet = workbook['discard']
        col_index = 1  

        # Retrieve all string values and store in a list
        # -----------------------------------------------
        column_values_list = [str(sheet.cell(row=i, column=col_index).value) \
                              for i in range(2, sheet.max_row + 1)]

        # Check data for erroneous input values
        # --------------------------------------
        bad_list = []
        good_list = []
        for value in column_values_list :
            value = ("'" + value + "'")
            mysql_select_query = ("SELECT * FROM inv WHERE code = " + value)
            cursor = CONNECTION.cursor(buffered = True)
            cursor.execute(mysql_select_query) 
            sel_res = cursor.fetchone()

            if sel_res is None:
                bad_list.append(value)

            else:
                good_list.append(value)

        # Update database
        # ---------------
        for i in good_list:
            sql_up_query = ("""UPDATE inv SET discard = 1 WHERE code = """ \
                             + i)
            cursor.execute(sql_up_query)
            CONNECTION.commit()  
        for i in bad_list:
            today = date.today()
            data = (i, "0", today)
            save_query = ("INSERT INTO badlist (value,correction_made, \
                           date_create)"
                        "VALUES (%s, %s, %s)")
            cursor = CONNECTION.cursor()
            cursor.execute(save_query, data)
            CONNECTION.commit()    
            
        # Print update messages
        # ----------------------
        print("Data Updated!") 
        print("These values were not found in the database")
        res = ", ".join(str(x) for x in bad_list)
        print(res)
        print('')

# ==============================================================================
#  main entry point for program
#  =============================================================================    
        
def main():
    """
    ============================================================================
    Function:       main()
    Purpose:        entry point to program
    Parameter(s):   -None-
    Return:         -None-
    ============================================================================
    """
    global CONNECTION

    menu()

    print("init_code_no: " + str(init_code_no))
    print("code_no: " + str(code_no))

    print("Closing Database Connection . . .")
    CONNECTION.close()
    print("bye . . .")

if __name__ == "__main__":
    main()

